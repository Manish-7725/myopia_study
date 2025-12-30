from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAdminUser

# API: Export all students with baseline and follow-up data (for frontend export)
@api_view(["GET"])
@permission_classes([IsAdminUser])
def students_export_json(request):
    students = Student.objects.all().order_by("-created_at")
    result = []
    for s in students:
        # Get latest baseline data for each section using correct related_name
        lifestyle = s.lifestyles.all().order_by("-created_at").first()
        environment = s.environments.all().order_by("-created_at").first()
        history = s.histories.all().order_by("-created_at").first()
        awareness = s.awareness.all().order_by("-created_at").first()
        ocular = s.ocular.all().order_by("-created_at").first()

        # All follow-ups for this student (with nested data)
        followups = []
        for f in s.followups.all().order_by("-created_at"):
            # Environmental, history, ocular are OneToOne related, may not exist
            env = getattr(f, "environmental", None)
            hist = getattr(f, "history", None)
            ocul = getattr(f, "ocular", None)
            followups.append({
                "id": f.id,
                "last_visit": str(f.last_visit) if f.last_visit else None,
                "next_visit": str(f.next_visit) if f.next_visit else None,
                "status": f.status,
                "notes": f.notes,
                "created_at": str(f.created_at),
                "environmental": {k.name: getattr(env, k.name) for k in env._meta.fields if k.name not in ("id", "created_at", "followup")} if env else None,
                "history": {k.name: getattr(hist, k.name) for k in hist._meta.fields if k.name not in ("id", "created_at", "followup")} if hist else None,
                "ocular": {k.name: getattr(ocul, k.name) for k in ocul._meta.fields if k.name not in ("id", "created_at", "followup")} if ocul else None,
            })

        result.append({
            "student_id": s.student_id,
            "name": s.name,
            "age": s.age,
            "gender": s.gender,
            "school_name": s.school_name,
            "created_at": str(s.created_at),
            "lifestyle": LifestyleBehaviorSerializer(lifestyle).data if lifestyle else None,
            "environment": EnvironmentalFactorSerializer(environment).data if environment else None,
            "history": ClinicalHistorySerializer(history).data if history else None,
            "awareness": AwarenessSafetySerializer(awareness).data if awareness else None,
            "ocular": OcularExaminationSerializer(ocular).data if ocular else None,
            "followups": followups,
        })
    return Response(result)
import csv
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from openpyxl import Workbook
from .serializers import AdminStudentListSerializer
from django.shortcuts import get_object_or_404
from rest_framework import status
from .serializers import StudentProfileSerializer
from django.db.models import Q



from .models import Student, ClinicalHistory, FollowUp
from .permissions import IsAdmin

from rest_framework.response import Response
from django.utils.timezone import now
from datetime import timedelta
from collections import defaultdict
from django.db.models import Count
from django.contrib.auth.models import User
from .models import FollowUp

from .models import (
    Student,
    ClinicalHistory,
    LifestyleBehavior,
    EnvironmentalFactor,
    AwarenessSafety,
    OcularExamination,
)

from .serializers import FollowUpSerializer, AdminUserSerializer, LifestyleBehaviorSerializer, EnvironmentalFactorSerializer, ClinicalHistorySerializer, AwarenessSafetySerializer, OcularExaminationSerializer, FollowUpCreateSerializer


@api_view(["GET", "PUT"])
@permission_classes([IsAdmin])
def admin_edit_visit(request, student_id, visit_date):
    student = get_object_or_404(Student, student_id=student_id)

    lifestyle = LifestyleBehavior.objects.get(student=student, visit_date=visit_date)
    environment = EnvironmentalFactor.objects.get(student=student, visit_date=visit_date)
    history = ClinicalHistory.objects.get(student=student, visit_date=visit_date)
    awareness = AwarenessSafety.objects.get(student=student, visit_date=visit_date)
    ocular = OcularExamination.objects.get(student=student, visit_date=visit_date)

    if request.method == "GET":
        return Response({
            "lifestyle": LifestyleBehaviorSerializer(lifestyle).data,
            "environment": EnvironmentalFactorSerializer(environment).data,
            "history": ClinicalHistorySerializer(history).data,
            "awareness": AwarenessSafetySerializer(awareness).data,
            "ocular": OcularExaminationSerializer(ocular).data,
        })

    if request.method == "PUT":
        LifestyleBehaviorSerializer(lifestyle, data=request.data["lifestyle"], partial=True).is_valid(raise_exception=True)
        EnvironmentalFactorSerializer(environment, data=request.data["environment"], partial=True).is_valid(raise_exception=True)
        ClinicalHistorySerializer(history, data=request.data["history"], partial=True).is_valid(raise_exception=True)
        AwarenessSafetySerializer(awareness, data=request.data["awareness"], partial=True).is_valid(raise_exception=True)
        OcularExaminationSerializer(ocular, data=request.data["ocular"], partial=True).is_valid(raise_exception=True)

        LifestyleBehaviorSerializer(lifestyle, data=request.data["lifestyle"], partial=True).save()
        EnvironmentalFactorSerializer(environment, data=request.data["environment"], partial=True).save()
        ClinicalHistorySerializer(history, data=request.data["history"], partial=True).save()
        AwarenessSafetySerializer(awareness, data=request.data["awareness"], partial=True).save()
        OcularExaminationSerializer(ocular, data=request.data["ocular"], partial=True).save()

        return Response({"message": "Visit updated successfully"})


@api_view(["GET"])
@permission_classes([IsAdmin])
def admin_overview(request):
    today = now().date()

    return Response({
        "total_students": Student.objects.count(),
        "total_followups": FollowUp.objects.count(),
        "due_this_week": FollowUp.objects.filter(
            status="Due",
            next_visit__gte=today,
            next_visit__lte=today + timedelta(days=7)
        ).count(),
        "missed_followups": FollowUp.objects.filter(status="Overdue").count(),
    })


@api_view(["GET"])
@permission_classes([IsAdmin])
def activity_analytics(request):
    """
    Provides data for the activity analytics chart.
    Returns the number of new students and follow-ups for the last 7 days.
    """
    today = now().date()
    seven_days_ago = today - timedelta(days=6)

    # Initialize data structure
    analytics_data = {
        "labels": [(seven_days_ago + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)],
        "students": [0] * 7,
        "followups": [0] * 7,
    }

    # Fetch student data
    student_counts = Student.objects.filter(
        created_at__date__gte=seven_days_ago
    ).values('created_at__date').annotate(count=Count('student_id'))

    # Fetch followup data
    followup_counts = FollowUp.objects.filter(
        created_at__date__gte=seven_days_ago
    ).values('created_at__date').annotate(count=Count('id'))


    # Populate analytics_data
    for entry in student_counts:
        try:
            idx = (entry['created_at__date'] - seven_days_ago).days
            if 0 <= idx < 7:
                analytics_data["students"][idx] = entry['count']
        except (IndexError, TypeError):
            pass

    for entry in followup_counts:
        try:
            idx = (entry['created_at__date'] - seven_days_ago).days
            if 0 <= idx < 7:
                analytics_data["followups"][idx] = entry['count']
        except (IndexError, TypeError):
            pass
            
    return Response(analytics_data)


from rest_framework.pagination import PageNumberPagination
from django.db.models import Q

class StudentPagination(PageNumberPagination):
    page_size = 50

from django.db.models import Q
from django.core.paginator import Paginator
from django.utils.dateparse import parse_date
from rest_framework.permissions import IsAuthenticated


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def admin_students(request):

    qs = Student.objects.all().order_by("-created_at")

    q = request.GET.get("q")
    if q:
        qs = qs.filter(
            Q(student_id__icontains=q) |
            Q(name__icontains=q) |
            Q(school_name__icontains=q)
        )

    from_date = parse_date(request.GET.get("from_date")) if request.GET.get("from_date") else None
    to_date = parse_date(request.GET.get("to_date")) if request.GET.get("to_date") else None
    gender = request.GET.get("gender")
    school = request.GET.get("school")

    if from_date:
        qs = qs.filter(created_at__date__gte=from_date)
    if to_date:
        qs = qs.filter(created_at__date__lte=to_date)
    if gender:
        qs = qs.filter(gender__iexact=gender)
    if school:
        qs = qs.filter(school_name__icontains=school)

    paginator = Paginator(qs, 10)
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)

    serializer = AdminStudentListSerializer(page_obj, many=True)

    return Response({
        "count": paginator.count,
        "next": page_obj.next_page_number() if page_obj.has_next() else None,
        "previous": page_obj.previous_page_number() if page_obj.has_previous() else None,
        "results": serializer.data
    })




@api_view(["GET", "PUT", "DELETE"])
@permission_classes([IsAdmin])
def admin_student_detail(request, student_id):
    student = get_object_or_404(Student, student_id=student_id)

    if request.method == "GET":
        serializer = StudentProfileSerializer(student)
        return Response(serializer.data)

    if request.method == "PUT":
        serializer = StudentProfileSerializer(
            student,
            data=request.data,
            partial=True
        )
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Student updated"})
        return Response(serializer.errors, status=400)

    if request.method == "DELETE":
        student.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(["GET"])
@permission_classes([IsAdmin])
def admin_followups(request):
    qs = ClinicalHistory.objects.select_related("student").order_by("-visit_date")[:50]
    return Response(FollowUpSerializer(qs, many=True).data)


@api_view(["GET"])
@permission_classes([IsAdmin])
def admin_users(request):
    return Response(
        AdminUserSerializer(User.objects.all(), many=True).data
    )


@api_view(["POST"])
@permission_classes([IsAdmin])
def admin_create_user(request):
    from django.contrib.auth.models import User

    username = request.data.get("username")
    password = request.data.get("password")
    email = request.data.get("email", "")
    role = request.data.get("role")  # Admin / Clinician / Surveyor

    if not username or not password:
        return Response(
            {"error": "username and password required"},
            status=status.HTTP_400_BAD_REQUEST
        )

    if User.objects.filter(username=username).exists():
        return Response(
            {"error": "username already exists"},
            status=status.HTTP_400_BAD_REQUEST
        )

    # ✅ CREATE USER (ONLY THIS MATTERS)
    user = User.objects.create_user(
        username=username,
        password=password,
        email=email
    )

    # ✅ ROLE MAPPING (ONLY ADMIN VS USER)
    if role == "Admin":
        user.is_staff = True
        user.is_superuser = True
    else:
        user.is_staff = False
        user.is_superuser = False

    user.save()

    return Response(
        {
            "message": "User created",
            "username": user.username,
            "role": "admin" if user.is_staff else "user"
        },
        status=status.HTTP_201_CREATED
    )



@api_view(["GET"])
@permission_classes([IsAdmin])
def export_students_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    ws.append(["Student ID", "Name", "Age", "Gender", "Created At"])

    for s in Student.objects.all().order_by("-created_at"):
        ws.append([
            s.student_id,
            s.name,
            s.age,
            s.gender,
            s.created_at.strftime("%Y-%m-%d"),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="students.xlsx"'
    wb.save(response)
    return response


@api_view(["GET"])
@permission_classes([IsAdmin])
def export_followups_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="followups.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "ID",
        "Student Name",
        "Visit Date",
        "Compliance",
        "Power Changed (3 yrs)"
    ])

    qs = ClinicalHistory.objects.select_related("student").order_by("-visit_date")

    for f in qs:
        writer.writerow([
            f.id,
            f.student.name,
            f.visit_date,
            f.compliance,
            "Yes" if f.power_changed_last_3yrs else "No",
        ])

    return response


@api_view(["GET"])
@permission_classes([IsAdmin])
def export_followups_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Follow-Ups"

    ws.append([
        "ID",
        "Student Name",
        "Visit Date",
        "Compliance",
        "Power Changed (3 yrs)"
    ])

    qs = ClinicalHistory.objects.select_related("student").order_by("-visit_date")

    for f in qs:
        ws.append([
            f.id,
            f.student.name,
            f.visit_date.strftime("%Y-%m-%d"),
            f.compliance,
            "Yes" if f.power_changed_last_3yrs else "No",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="followups.xlsx"'
    wb.save(response)
    return response


from django.utils.dateparse import parse_date

# @api_view(["POST"])
# @permission_classes([IsAdmin])
# def export_students_csv(request):
#     data = request.data

#     search = data.get("search", "").strip()
#     ids = data.get("ids", [])
#     from_date = parse_date(data.get("from_date")) if data.get("from_date") else None
#     to_date = parse_date(data.get("to_date")) if data.get("to_date") else None
#     export_format = data.get("format", "csv")

#     qs = Student.objects.all().order_by("-created_at")

#     # 🔍 Search filter
#     if search:
#         qs = qs.filter(
#             Q(student_id__icontains=search) |
#             Q(name__icontains=search)
#         )

#     # ☑ Selected rows
#     if ids:
#         qs = qs.filter(student_id__in=ids)

#     # 📅 Date range
#     if from_date:
#         qs = qs.filter(created_at__date__gte=from_date)
#     if to_date:
#         qs = qs.filter(created_at__date__lte=to_date)

#     # ================= CSV =================
#     if export_format == "csv":
#         response = HttpResponse(content_type="text/csv")
#         response["Content-Disposition"] = 'attachment; filename="students.csv"'

#         writer = csv.writer(response)
#         writer.writerow(["Student ID", "Name", "Age", "Gender", "Created At"])

#         for s in qs:
#             writer.writerow([
#                 s.student_id,
#                 s.name,
#                 s.age,
#                 s.gender,
#                 s.created_at.date()
#             ])

#         return response

#     # ================= EXCEL =================
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Students"

#     ws.append(["Student ID", "Name", "Age", "Gender", "Created At"])

#     for s in qs:
#         ws.append([
#             s.student_id,
#             s.name,
#             s.age,
#             s.gender,
#             s.created_at.date()
#         ])

#     response = HttpResponse(
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     response["Content-Disposition"] = 'attachment; filename="students.xlsx"'
#     wb.save(response)
#     return response

@api_view(["POST"])
@permission_classes([IsAdmin])
def create_followup(request):
    serializer = FollowUpCreateSerializer(data=request.data)
    if serializer.is_valid(raise_exception=True):
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


from rest_framework.pagination import PageNumberPagination
from django.db.models import Q

class StudentPagination(PageNumberPagination):
    page_size = 50

